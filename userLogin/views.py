import base64
import os, requests
from django.shortcuts import render,redirect,HttpResponse
from django.core.mail import send_mail
from django.conf import settings
from django.contrib.auth.hashers import make_password,check_password
from django.contrib.auth import login as auth_login,logout as auth_logout,authenticate
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.urls import reverse
from .models import User,Notice,CharityAlert,DonatedTo,DonationTransactionID,TrustDonationID,NGOMedia,NGOPost
import random, qrcode, uuid, razorpay,qrcode
from io import BytesIO
from django.utils import timezone
from datetime import datetime,timedelta
from django.templatetags.static import static
import io
from PIL import Image
import openpyxl
from openpyxl import Workbook
from django_ratelimit.decorators import ratelimit
from django.views.decorators.csrf import csrf_exempt
key_id = settings.RAZORPAY_KEY_ID
key_secret = settings.RAZORPAY_KEY_SECRET

def generate_qr_code(upi_id, amount, beneficiary_name):
    upi_url = f"upi://pay?pa={upi_id}&pn={beneficiary_name}&am={amount}&cu=INR&tn=NGO+Support"
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(upi_url)
    qr.make(fit=True)
    img_io = BytesIO()
    qr.make_image().save(img_io, 'PNG')
    img_io.seek(0)
    img_base64 = base64.b64encode(img_io.getvalue()).decode('ascii')
    return img_base64



# Create your views here.
def login(request):
    if request.method=="POST":
        userId = request.POST.get("userID")
        password = request.POST.get("Password")
        try:
            user=None
            if userId.isdigit():
                user = User.objects.filter(mobile=userId).first()
            else:
                user = User.objects.filter(email=userId).first()

            if user and check_password(password,user.password):
                user=authenticate(request, username=userId,password=password)
                if user:
                    auth_login(request,user)
                    return redirect("profile")
                else:
                    return render(request,"loginPage.html",{"PasswordInvalid": "Password is invalid"})
                
            else:
                if user is None:
                    return render(request,"loginPage.html",{"EmailInvalid": "Email is invalid"})
                else:
                    return render(request,"loginPage.html",{"PasswordInvalid": "Password is invalid"})
            
        except Exception as e:
            return render(request,"errorPage.html",{"error": e})
    else:
        return render(request,"loginPage.html")
    
@login_required
def logout(request):
    auth_logout(request)  
    return redirect('home') 


def forgetPassword(request):
    try: 
        if request.method=="POST":
            email = request.POST.get("email")
            if User.objects.filter(email=email).exists():
                otp = str(random.randint(1000,9999))
                token = uuid.uuid4().hex
                send_mail(
                    "Unique Group Of Mankind Trust - Password Reset OTP",
                    f"""
Dear User,

We are pleased to inform you that your password reset request has been processed. Please find your OTP below:

{otp}

This OTP is valid for a short period. If you did not request an email verification, please ignore this message.

To proceed with the verification, kindly enter the OTP in the designated field on the verification page. If you encounter any issues, please contact our support team for assistance.

Thank you for choosing Unique Group Of Mankind Trust.

Best regards,
The Unique Group Of Mankind Trust Team
""",
                    settings.EMAIL_HOST_USER,
                    [email],fail_silently=False)
                request.session["email"] = email
                request.session['otp'] = otp
                request.session['token'] = token
                return redirect(reverse("otp",kwargs={"email":email,"token":token}))
            else:
                return render(request,"forgetPage.html",{"notfound": "Email not found","title":"Forget Password"})
        else:
            return render(request,"forgetPage.html",{"title":"Forget Password"})
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})
    
def verify_otp(request,email,token):
    try:
        if request.session.get("token")==token:
            if request.method == "POST" and request.session.get("token")==token:
                otpVerify = request.POST.get("otpVerify")
                # Get the OTP value from the session
                session_otp = request.session.get("otp")
                if otpVerify == session_otp:
                    del request.session["token"]
                    del request.session["otp"]
                    del session_otp
                    token = uuid.uuid4().hex
                    request.session["email"] = request.session.get("email")
                    request.session["token!"] = token
                    # OTP is valid, proceed to reset password or perform other actions
                    return redirect(reverse("reset_password",kwargs={"token":token}))
                else:
                    return render(request, "OTPpage.html", {"error": "Invalid OTP","title":"Forget Password"})
            else:
                return render(request, "OTPpage.html",{"title":"Forget Password"})
        return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

def resetPassword(request,token):
    try:
        if request.session.get("token!")==token:
            if request.method=="POST" and request.session.get("token!")==token:
                del request.session["token!"]
                password = request.POST.get("resetPassword")
                email = request.session.get("email")
                hashPassword = make_password(password)
                User.objects.filter(email=email).update(password=hashPassword)
                del email
                del request.session["email"]
                return render(request,"passwordChanged.html")
            else:
                return render(request, "password_reset_confirm.html")
        return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})
        

@csrf_exempt
def sendOTP(request):
    try:
        if request.method=="POST":
            email = request.POST.get("email")
            if User.objects.filter(email=email).exists():
                return render(request,"forgetPage.html",{"notfound": "Email already Taken","title":"Email Verify"})
            else:
                otp = str(random.randint(1000,9999))
                token = uuid.uuid4().hex
                send_mail(
                    "Unique Group Of Mankind Trust - Email Verification OTP",
                    f"""
Dear User,

We are pleased to inform you that your email verification request has been processed. Please find your OTP below:

{otp}

This OTP is valid for a short period. If you did not request an email verification, please ignore this message.

To proceed with the verification, kindly enter the OTP in the designated field on the verification page. If you encounter any issues, please contact our support team for assistance.

Thank you for choosing Unique Group Of Mankind Trust.

Best regards,
The Unique Group Of Mankind Trust Team
""",
                    settings.EMAIL_HOST_USER,
                    [email],fail_silently=False)
                request.session["email"] = email
                request.session['otp'] = otp
                request.session['token'] = token
                return redirect(reverse("VerifyEmail",kwargs={"email":email,"token":token}))
        else:
            return render(request,"forgetPage.html",{"title":"Email Verify"})
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})
    
def verify_email(request,email,token):
    try:
        if request.session.get("token")==token:
            if request.method == "POST" and request.session.get("token")==token:
                otpVerify = request.POST.get("otpVerify")
                # Get the OTP value from the session
                session_otp = request.session.get("otp")
                if otpVerify == session_otp:
                    del request.session["token"]
                    del request.session["otp"]
                    del session_otp
                    token = uuid.uuid4().hex
                    # OTP is valid, proceed to reset password or perform other actions
                    request.session["token!"] = token
                    return redirect(reverse("register",kwargs={"token":token}))
                else:
                    return render(request, "OTPpage.html", {"error": "Invalid OTP","title":"Email Verify"})
            else:
                return render(request, "OTPpage.html",{"title":"Email Verify"})
        return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})
    
# ---------------       P A Y M E N T
def register(request,token):
    try:
        if request.session.get("token!")==token:
            if request.method=="POST" and request.session.get("token!")==token:
                name=request.POST.get("Name")
                email=request.session.get("email")
                dob = request.POST.get("DOB")
                gender = request.POST.get("Gender")
                blood_group = request.POST.get("BG")
                state = request.POST.get("State")
                district = request.POST.get("District")
                zipCode = request.POST.get("ZipCode")
                nearBy = request.POST.get("nearBy")
                address = request.POST.get("Address")
                mobile = request.POST.get("MobileNumber")
                aadhar = request.POST.get("Aadhar")
                aadharPicture = request.FILES.get("aadharPicture")
                password = request.POST.get("Password")
                # Nominee1-Details
                nominee1_name=request.POST.get("Nominee1Name")
                nominee1_mobile=request.POST.get("Nominee1Mobile")
                nominee1_relation=request.POST.get("Nominee1Relation")
                # Chronic Disease
                disease = request.POST.get("flexRadioDefault")
                # Referral-Details
                referral_id=request.POST.get("ReferralID")
                if referral_id:
                    if not (referral_id.startswith("UGMT") and len(referral_id)==7):
                        return render(request,"registerPage.html",{"ReferralidError": "UGMT ID doesn't exist"})
                    referral_id = referral_id[4:]
                    referral_id = str(int(referral_id))
                    if not User.objects.filter(id=str(referral_id)).exists():
                        return render(request,"registerPage.html",{"ReferralidError": "UGMT ID doesn't exist"})

                # check if email and aadhar are already taken
                if User.objects.filter(email=email).exists():
                    return render(request,"registerPage.html",{"EmailError": "Email already taken"})
                elif User.objects.filter(aadhar=aadhar).exists():
                    return render(request,"registerPage.html",{"AadharError": "Aadhar Number already exist"})
                elif User.objects.filter(mobile=str(mobile)).exists():
                    return render(request,"registerPage.html",{"MobileError": "Phone Number already exist"})
                else:
                    # PAYMENT GATEWAY
                    # Create a Razorpay order
                    client = razorpay.Client(auth=(key_id,key_secret))
                    fee_amount = 20000  # amount in rupees
                    # fee_amount = 100  # amount in rupees
                    fee_currency = 'INR'
                    fee_receipt = 'UGMT Member fees'

                    response = client.order.create(
                        {
                        "amount":fee_amount,
                        "currency":fee_currency,
                        "receipt":fee_receipt,
                        "payment_capture":'1'
                        }
                    )

                    fee_id = response['id']
                    fee_status = response['status']

                    # If fee is created successfully, redirect to payment page
                    if fee_status == 'created':
                        hashPassword = make_password(password)
                        request.session["USER"] = dict(
                            name=name,
                            email=email,
                            dob=dob,
                            gender=gender,
                            blood_group=blood_group,
                            state=state,
                            district=district,
                            zip_code = zipCode,
                            near_by = nearBy,
                            address=address,
                            mobile=mobile,
                            aadhar=aadhar,
                            password=hashPassword,
                            disease = disease,
                            # nominee1
                            nominee1_name=nominee1_name,
                            nominee1_mobile=nominee1_mobile,
                            nominee1_relation=nominee1_relation,
                        )
                        if referral_id:
                            request.session["USER"]["referral_id"] = referral_id
                        if aadharPicture:
                            aadhar_picture_bytes = aadharPicture.read()
                            aadhar_picture_base64 = base64.b64encode(aadhar_picture_bytes)
                            request.session["AADHAR"] = aadhar_picture_base64.decode('ascii')
                        else:
                            request.session["AADHAR"]  = None
                        context = {
                            'razorpay_key': key_id,
                            'order_id': fee_id,
                            'amount': fee_amount,
                            "mobile":mobile,
                            "email":email,
                            'currency': 'INR',
                            'image': static("images/logonavbar.gif"),
                        }
                        return render(request, 'payment/membership_Payment.html', context)
                    else:
                        return render(request, "payment/paymentFail.html",{"error": "Status Failed"})
            else:
                return render(request,"registerPage.html")
        return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

# ---------------------     P A Y M E N T   &   R E G I S T R A T I O N 
def membershipPayment(request):
    try:
        order_id = request.POST.get("razorpay_order_id")
        payment_id = request.POST.get("razorpay_payment_id")
        signature = request.POST.get("razorpay_signature")

        # Verifying the payment signature
        params_dict = {
            'razorpay_order_id': order_id,
            'razorpay_payment_id': payment_id,
            'razorpay_signature': signature
        }
        del request.session["token!"]

        client = razorpay.Client(auth=(key_id,key_secret))    
        result = client.utility.verify_payment_signature(params_dict)
        if result:
            user_data = request.session.get('USER')
            aadharPicture = request.session.get("AADHAR")
            if aadharPicture:
                aadhar_picture_base64 = aadharPicture
                aadhar_picture_bytes = base64.b64decode(aadhar_picture_base64)
                aadharPic = InMemoryUploadedFile(
                    BytesIO(aadhar_picture_bytes),
                    'file',
                    'aadhar_picture.jpg',
                    'image/jpeg',
                    len(aadhar_picture_bytes),
                    None
                )
                user = User(
                    name=user_data['name'],
                    email=user_data['email'],
                    dob=user_data['dob'],
                    gender=user_data['gender'],
                    blood_group=user_data['blood_group'],
                    state=user_data['state'],
                    district=user_data['district'],
                    zip_code=user_data['zip_code'],
                    near_by=user_data['near_by'],
                    address=user_data['address'],
                    mobile=user_data['mobile'],
                    aadhar=user_data['aadhar'],
                    aadharPhoto=aadharPic,
                    password=user_data['password'],
                    chronic_disease = user_data["disease"],
                    registerDate = timezone.now(),                    
                    expiring_Date = (timezone.now()+timedelta(days=365)),
                    # NOMINEE-1
                    nominee1_name = user_data['nominee1_name'],
                    nominee1_mobile = user_data['nominee1_mobile'],
                    nominee1_relation = user_data['nominee1_relation'],
                )
            else:
                user = User(
                    name=user_data['name'],
                    email=user_data['email'],
                    dob=user_data['dob'],
                    gender=user_data['gender'],
                    blood_group=user_data['blood_group'],
                    state=user_data['state'],
                    district=user_data['district'],
                    zip_code=user_data['zip_code'],
                    near_by=user_data['near_by'],
                    address=user_data['address'],
                    mobile=user_data['mobile'],
                    aadhar=user_data['aadhar'],
                    password=user_data['password'],
                    registerDate = timezone.now(),
                    chronic_disease = user_data["disease"],
                    expiring_Date = (timezone.now()+timedelta(days=365)),
                    # NOMINEE-1
                    nominee1_name = user_data['nominee1_name'],
                    nominee1_mobile = user_data['nominee1_mobile'],
                    nominee1_relation = user_data['nominee1_relation'],
                )
            if 'referral_id' in user_data:
                user.refer_name = User.objects.filter(id=str(user_data['referral_id'])).first().name
                user.refer_mobile = User.objects.filter(id=str(user_data['referral_id'])).first().mobile
                user.refer_id = user_data['referral_id']
            user.save()
            send_mail(
                "New User Registerd",
                f"Name: {user_data['name']}\n \n Number: {user_data['mobile']}\n Email: {user_data['email']}\nAadhar Number: \n{user_data['aadhar']}",
                settings.EMAIL_HOST_USER,['info@ugmt.org'],fail_silently=False)
            del request.session['USER']
            del request.session['AADHAR']
            return render(request,"payment/paymentSuccess.html")
        else:
            return render(request, "payment/paymentFail.html",{"error":"Contact Administrative, payment gateway issue arises"})

    except Exception as e:
        return render(request,"errorPage.html",{"error": e})
    
# ---------------------     C H A R I T Y  -  P A Y M E N T
@login_required
def CharityPayment(request,token,nominee_id,token1,charityId):
    try:
        if request.session.get("TOKENstatement")==token:
            nominee = User.objects.filter(id=nominee_id).first()
            charity = CharityAlert.objects.filter(id=charityId).first()
            qr_code_image = generate_qr_code(charity.upi_id, charity.amount_per_member, nominee.name)
            if request.method=="POST":
                transactionID = request.POST.get("transactionID")
                if not DonationTransactionID.objects.filter(transaction_id=transactionID).first():
                    del request.session["TOKENstatement"]
                    donation_transaction_details = DonationTransactionID.objects.create(
                        transaction_id=transactionID,
                        user=request.user,
                        charity_alert = CharityAlert.objects.filter(id=charityId).first(),
                    )
                    donated_to = DonatedTo.objects.create(
                        user=request.user,
                        charity_alert = CharityAlert.objects.filter(id=charityId).first()
                    )
                    donation_transaction_details.save()
                    donated_to.save()
                    return redirect("donatedTo")
                else:
                    return render(request, "payment/DonatePayment.html", {"UserInfo":nominee,"CharityInfo":charity,"QRcode":qr_code_image,"error": "Transaction ID already exists"})
            else:
                return render(request, "payment/DonatePayment.html",{"UserInfo":nominee,"CharityInfo":charity,"QRcode":qr_code_image})
        else:
            return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

@login_required
def downloadQR(request, token,nomineeID,charityID):
    try:
        if request.session.get("TOKENstatement")==token:
            nominee = User.objects.filter(id=nomineeID).first()
            charity = CharityAlert.objects.filter(id=charityID).first()
            qr_code_image = generate_qr_code(charity.upi_id, charity.amount_per_member, nominee.name)
            # Check if qr_code_image is a valid base64-encoded string
            try:
                image_data = base64.b64decode(qr_code_image)
            except base64.binascii.Error:
                return render(request, "errorPage.html", {"error": "Invalid QR code"})
            # Check if image_data is a valid image
            try:
                image = Image.open(io.BytesIO(image_data))
            except IOError:
                return render(request, "errorPage.html", {"error": "Invalid image"})
            response = HttpResponse(content_type='image/png')
            response['Content-Disposition'] = 'attachment; filename="qrcode.png"'
            image.save(response, 'PNG')
            return response
        else:
            return render(request, "errorPage.html", {"error": "Unable to Download QR"})
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

# -----------   R E N E W   P A Y M E N T
@login_required
def renew_membership(request,token):
    try:
        if token==request.session.get("RENEW_TOKEN"):
            del request.session["RENEW_TOKEN"]
            # Create a new payment order
            client = razorpay.Client(auth=(key_id, key_secret))
            order = client.order.create(
                {
                "amount": 20000,  # amount in rupee
                "currency": "INR",
                "receipt": "Membership Renewal",
                "payment_capture": "0"
                }
            )
            status = order['status']
            if status=="created":
                token = uuid.uuid4().hex
                request.session["token"] = token
                context = {
                    'razorpay_key': key_id,
                    'order_id': order['id'],
                    'amount': 100,
                    'currency': 'INR',
                    'image': static("images/logonavbar.gif"),
                    'token':token,
                }
                return render(request, 'payment/renewMembership.html', context)
            else:
                return render(request, "payment/paymentFail.html",{"error": "Status Failed"})
        else:
            return render(request, "errorPage.html", {"error": "error in renewMembership"})
    except Exception as e:
        return render(request, "errorPage.html", {"error": e})

@login_required
def renew_membership_success(request,token):
    try:
        if token==request.session.get("token"):
            del request.session["token"]
            order_id = request.POST.get("razorpay_order_id")
            payment_id = request.POST.get("razorpay_payment_id")
            signature = request.POST.get("razorpay_signature")

            # Verifying the payment signature
            params_dict = {
                'razorpay_order_id': order_id,
                'razorpay_payment_id': payment_id,
                'razorpay_signature': signature
            }

            client = razorpay.Client(auth=(key_id, key_secret))
            result = client.utility.verify_payment_signature(params_dict)

            if result:

                capture=client.payment.capture(payment_id,100)
                if capture["captured"]:
                # Update the user's membership status
                    user = request.user
                    user.expiring_Date = datetime.now().date() + timedelta(days=365)
                    user.save()
                    return redirect("profile")
                else:
                    return render(request, "payment/paymentFail.html", {"error": "Payment verification failed"})
            else:
                return render(request, "payment/paymentFail.html", {"error": "Payment verification failed"})
        else:
            return render(request, "payment/paymentFail.html", {"error": "Payment verification TOKEN failed"})
    except Exception as e:
        return render(request, "errorPage.html", {"error": e})

# ---------------------     L O G I N - R E Q U I R E D
    #------------------     D O W N L O A D   I N F O
@login_required
def download_users_data(request):
    # Get the search query parameters
    search,filtered="","" 
    search_input = request.GET.get('searchInput', '')
    filtered = request.GET.get("filtered", '')

    # Initialize the queryset
    queryset = User.objects.all()
    # Apply the search filter if any
    if search_input and filtered:
        if filtered == 'SN' and search_input.isdigit():
            queryset = queryset.filter(id__icontains=search_input)
        elif filtered == 'Name':
            queryset = queryset.filter(name__icontains=search_input)
        elif filtered == 'Email':
            queryset = queryset.filter(email__icontains=search_input)
        elif filtered == 'PhN' and search_input.isdigit():
            queryset = queryset.filter(mobile__icontains=search_input)
        else:
            queryset = User.objects.all()

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Set the header row
    headers = ['UGMT ID', 'Name', 'Email', 'DOB','Designation', 'Verified','Gender', 'Blood Group',
                'State', 'District','Address', 'Zip Code','Near By','Membership Expiring Date',
                'Mobile Number','Mobile Number Home', 'Aadhar Number', 'Register Date',
                "Profile Picture","Aadhaar Picture",'Referral Name','Referral Phone Number','Referral UGMT id',
                'Nominee Name','Nominee Phone Number','Nominee Relation']
    for i, header in enumerate(headers):
        ws.cell(row=1, column=i+1, value=header)

    # Iterate over the queryset and write the data to the worksheet
    def format_ugmt_id(user_id):
        user_id_str = str(user_id)
        if len(user_id_str) == 1:
            return f"UGMT00{user_id_str}"
        elif len(user_id_str) == 2:
            return f"UGMT0{user_id_str}"
        else:
            return f"UGMT{user_id_str}"
        
    for i, user in enumerate(queryset):
        ugmtid = format_ugmt_id(user.id)
        ws.cell(row=i+2, column=1, value=ugmtid)
        ws.cell(row=i+2, column=2, value=user.name)
        ws.cell(row=i+2, column=3, value=user.email)
        ws.cell(row=i+2, column=4, value=user.dob)
        ws.cell(row=i+2, column=5, value=user.designation)
        ws.cell(row=i+2, column=6, value=user.is_verified)
        ws.cell(row=i+2, column=7, value=user.gender)
        ws.cell(row=i+2, column=8, value=user.blood_group)
        ws.cell(row=i+2, column=9, value=user.state)
        ws.cell(row=i+2, column=10, value=user.district)
        ws.cell(row=i+2, column=11, value=user.address)
        ws.cell(row=i+2, column=12, value=user.zip_code)
        ws.cell(row=i+2, column=13, value=user.near_by)
        ws.cell(row=i+2, column=14, value=user.expiring_Date.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=i+2, column=15, value=user.mobile)
        ws.cell(row=i+2, column=16, value=user.mobile_home)
        ws.cell(row=i+2, column=17, value=user.aadhar)
        ws.cell(row=i+2, column=18, value=user.registerDate)
        ws.cell(row=i+2, column=19, value=user.profilePic.url)
        ws.cell(row=i+2, column=20, value=user.aadharPhoto.url)
        ws.cell(row=i+2, column=21, value=user.refer_name)
        ws.cell(row=i+2, column=22, value=user.refer_mobile)
        ws.cell(row=i+2, column=23, value=user.refer_id)
        ws.cell(row=i+2, column=24, value=user.nominee1_name)
        ws.cell(row=i+2, column=25, value=user.nominee1_mobile)
        ws.cell(row=i+2, column=26, value=user.nominee1_relation)

    # Save the workbook to a file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="user_data.xlsx"'
    wb.save(response)
    return response


@login_required
def download_user_data(request,memberID):
    queryset = User.objects.filter(id=memberID)

    # Create a workbook and a worksheet
    wb = Workbook()
    ws = wb.active

    # Set the header row
    headers = ['UGMT id', 'Name', 'Email', 'DOB','Designation', 'Verified','Gender', 'Blood Group',
                'State', 'District','Address', 'Zip Code','Near By','Membership Expiring Date',
                'Mobile Number','Mobile Number Home', 'Aadhaar Number', 'Register Date',
                'Profile Picture','Aadhaar Picture','Referral Name','Referral Phone Number','Referral UGMT id',
                'Nominee-1 Name','Nominee-1 Phone Number','Nominee-1 Relation',]
    for i, header in enumerate(headers):
        ws.cell(row=1, column=i+1, value=header)

    # Iterate over the queryset and write the data to the worksheet
    def format_ugmt_id(user_id):
        user_id_str = str(user_id)
        if len(user_id_str) == 1:
            return f"UGMT00{user_id_str}"
        elif len(user_id_str) == 2:
            return f"UGMT0{user_id_str}"
        else:
            return f"UGMT{user_id_str}"
    for i, user in enumerate(queryset):
        ugmtid = format_ugmt_id(user.id)
        ws.cell(row=i+2, column=1, value=ugmtid)
        ws.cell(row=i+2, column=2, value=user.name)
        ws.cell(row=i+2, column=3, value=user.email)
        ws.cell(row=i+2, column=4, value=user.dob)
        ws.cell(row=i+2, column=5, value=user.designation)
        ws.cell(row=i+2, column=6, value=user.is_verified)
        ws.cell(row=i+2, column=7, value=user.gender)
        ws.cell(row=i+2, column=8, value=user.blood_group)
        ws.cell(row=i+2, column=9, value=user.state)
        ws.cell(row=i+2, column=10, value=user.district)
        ws.cell(row=i+2, column=11, value=user.address)
        ws.cell(row=i+2, column=12, value=user.zip_code)
        ws.cell(row=i+2, column=13, value=user.near_by)
        ws.cell(row=i+2, column=14, value=user.expiring_Date.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=i+2, column=15, value=user.mobile)
        ws.cell(row=i+2, column=16, value=user.mobile_home)
        ws.cell(row=i+2, column=17, value=user.aadhar)
        ws.cell(row=i+2, column=18, value=user.registerDate)
        ws.cell(row=i+2, column=19, value=user.profilePic.url)
        ws.cell(row=i+2, column=20, value=user.aadharPhoto.url)
        ws.cell(row=i+2, column=21, value=user.refer_name)
        ws.cell(row=i+2, column=22, value=user.refer_mobile)
        ws.cell(row=i+2, column=23, value=user.refer_id)
        ws.cell(row=i+2, column=24, value=user.nominee1_name)
        ws.cell(row=i+2, column=25, value=user.nominee1_mobile)
        ws.cell(row=i+2, column=26, value=user.nominee1_relation)

    # Save the workbook to a file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{user.name}-{user.id}.xlsx"'
    wb.save(response)
    return response


@login_required
def searchMember(request):
    try:  
        search = request.GET.get("searchInput", "") 
        filtered = request.GET.get("filtered", "")
        total_users = ""
        users = User.objects.none()
        if request.method=="POST": 
            search = request.POST.get("searchInput")
            filtered = request.POST.get("filtered")
        ugmtid = search
        if search and filtered:
            if filtered=="SN" and ugmtid.startswith("UGMT"):
                if len(ugmtid)>=7:
                    ugmtid = ugmtid[4:]
                    ugmtid = int(ugmtid)
                    users = User.objects.filter(id=ugmtid,is_verified=True)[::-1]
            elif filtered=="Name":
                users = User.objects.filter(name__icontains=search,is_verified=True)[::-1]
            elif filtered=="Email":
                users = User.objects.filter(email__icontains=search,is_verified=True)[::-1]
            elif filtered=="PhN" and search.isdigit():
                users = User.objects.filter(mobile__icontains=search,is_verified=True)[::-1]
            total_users = len(users)
        else:
            users = User.objects.filter(is_verified=True).order_by("id")
            total_users = len(users)
        paginator = Paginator(users, 15) 
        page_number = request.GET.get('page') 
        page_obj = paginator.get_page(page_number)
        return render(request,"userPanel/membersPanel.html",{
            "userData":page_obj,"searchValue":search,"filtered":filtered,"totalUser":total_users
            })
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

@login_required
def memberVerification(request):
    try:
        if request.user.is_superuser:
            if request.method=="POST":
                aadhar_number = request.POST.get("aadharNumber")
                User.objects.filter(aadhar=aadhar_number,is_verified=False).update(is_verified=True)
                unverify_members = User.objects.filter(is_verified=False).order_by("id")
                return render(request,"userPanel/verifyMembers.html",{"unverifyData":unverify_members})
            else:
                unverify_members = User.objects.filter(is_verified=False).order_by("id")
                return render(request,"userPanel/verifyMembers.html",{"unverifyData":unverify_members})
        else:
            return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

@login_required
def memberInfo(request,memberID):
    try:
        if request.user.is_superuser:
            member = User.objects.get(id=memberID)
            return render(request,"userPanel/membersInfo.html",{"memberData":member})
        else:
            return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e}) 

@login_required
def donatedTo(request):
    try:
        donatedTo = DonatedTo.objects.filter(user=request.user).order_by("id")[::-1]
        donationTranID = DonationTransactionID.objects.filter(user=request.user,is_matched=True)
        return render(request,"userPanel/donatedTo.html",{"donatedTO":donatedTo,"TransID":donationTranID})
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

@login_required
def profile(request):
    try:
        email = request.user
        if User.objects.filter(email=email).exists():
            userData = User.objects.get(email=email)
            registerDate=userData.registerDate
            year = userData.expiring_Date
            expire=True if year<=datetime.now().date() else False
            if request.method=="POST":
                print(request.POST)
                if "profile_picture_form" in request.POST:
                    profile_pic = request.FILES.get("profilePicture")
                    if profile_pic:                
                        if userData.profilePic != "defaultProfile.png":
                            userData.profilePic.delete()
                        userData.profilePic = profile_pic
                        userData.save()
                elif "aadhaar_picture_form" in request.POST:
                    aadhar_card = request.FILES.get("aadhaarPicture")   
                    if aadhar_card:
                        userData.aadharPhoto = aadhar_card
                        userData.save()
                return redirect("profile")
            else:
                if request.user.is_superuser:
                    membership = "SUPERUSER"
                else:
                    membership = ""
                return render(request,"userPanel/profilePage.html",{"userData":userData,"expire":expire})
    except Exception as e:
        return render(request,"errorPage.html",{"error": f"Error in profile {e}"})

@login_required
def renew_token(request):
    try:
        user = request.user
        if user.expiring_Date<=datetime.now().date():
            token = uuid.uuid4().hex
            request.session["RENEW_TOKEN"] = token
            return redirect(reverse("renewMembership",kwargs={"token":token}))
        else:
            return render(request,"errorPage.html",{"error": f"Error in TOKEN - PROFILE"})    
    except Exception as e:
        return render(request,"errorPage.html",{"error": f"Error in profile {e}"})


@login_required
def viewNotice(request):
    try:
        if request.method=="POST":
            toDelete = request.POST.getlist("checkBox")
            Notice.objects.filter(id__in=toDelete).update(is_delete=True)
            return redirect("viewNotice")
        notice = Notice.objects.filter(is_delete=False).order_by("id")[::-1]
        return render(request,"userPanel/viewNotice.html",{"Notices":notice})
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

@login_required
def viewAlert(request):
    try:
        if request.method=="POST":
            if request.POST.get("Delete")=="deleteButton":
                toDelete = request.POST.getlist("checkBox")
                CharityAlert.objects.filter(id__in=toDelete).update(is_delete=True)
            if request.POST.get("Delete")=="InactiveButton":
                changeStatus = request.POST.getlist("checkBox")
                if changeStatus:
                    current_status = CharityAlert.objects.filter(id__in=changeStatus).first().donation_status
                    CharityAlert.objects.filter(id__in=changeStatus).update(donation_status=True if current_status==False else False)
            return redirect("viewAlert")
        if request.user.is_superuser:
            charity_alert = CharityAlert.objects.filter(is_delete=False).order_by("id")[::-1]
        else:
            donated_to = DonatedTo.objects.filter(user=request.user).values('charity_alert')
            charity_alert = CharityAlert.objects.filter(is_delete=False, donation_status=False).exclude(id__in=donated_to).order_by("id")[::-1]
        return render(request,"userPanel/viewAlert.html",{"charityAlert":charity_alert})
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})


@login_required
def charityToken(request,id,charityId):
    try:
        token = uuid.uuid4().hex
        request.session["TOKENstatement"] = token
        return redirect(reverse("CharityPayment",kwargs={"token":token,"nominee_id":id,"token1":token,"charityId":charityId}))
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})
    

@login_required
def idCheckToken(request,charityId):
    try:
        token = uuid.uuid4().hex
        request.session["TOKENstatement"] = token
        return redirect(reverse("idCheck",kwargs={"token":token,"charityId":charityId,"token1":token}))
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})


@login_required
def checkID(request,token,charityId,token1):
    try:
        if request.session.get("TOKENstatement")==token:
            if request.method=="POST":
                id = request.POST.get("transactionIDs")
                transactionID = DonationTransactionID.objects.filter(transaction_id=id).first()
                if transactionID:
                    transactionID.is_matched=True
                    transactionID.save()
                else:
                    transactionIDs = DonationTransactionID.objects.filter(charity_alert=CharityAlert.objects.filter(id=charityId).first()).order_by("id")[::-1]
                    return render(request,"payment/checkTransactionIDs.html",{"transactionIDs":transactionIDs,"Error":"No transactionID found"})
            transactionIDs = DonationTransactionID.objects.filter(charity_alert=CharityAlert.objects.filter(id=charityId).first()).order_by("id")[::-1]
            return render(request,"payment/checkTransactionIDs.html",{"transactionIDs":transactionIDs})
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})


@login_required
def checkTrustID(request):
    try:
        if request.user.is_superuser:
            if request.method=="POST":
                id = request.POST.get("transactionIDs")
                transactionID = TrustDonationID.objects.filter(transaction_id=id).first()
                if transactionID:
                    transactionID.is_matched=True
                    transactionID.save()
                else:
                    transactionIDs = TrustDonationID.objects.all().order_by('is_matched')
                    return render(request,"payment/checkTrustTransactionID.html",{"transactionIDs":transactionIDs,"Error":"No transactionID found"})
            transactionIDs = TrustDonationID.objects.all().order_by('is_matched')
            return render(request,"payment/checkTrustTransactionID.html",{"transactionIDs":transactionIDs})
        else:
            return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

@login_required
def genNotice(request):
    try:
        if request.user.is_superuser:
            if request.method=="POST":
                notice = request.POST.get("noticeText")
                noticeObj = Notice.objects.create(
                    notice_text=notice,
                    date = datetime.now().date(),
                    created_by = request.user,
                    )
                noticeObj.save()
                return redirect("viewNotice")
            else:
                return render(request,"userPanel/genNotice.html")
        else:
            return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})

@login_required
def genAlert(request):
    try:
        if request.user.is_superuser:
            if request.method=="POST":
                alert = request.POST.get("alertText")
                receiverAadhar = request.POST.get("receiverAadhar")
                nominee_aadhar = request.POST.get("nomineeAadhar")
                bank = request.POST.get("bankName")
                upi = request.POST.get("upiId")
                totalAmount = request.POST.get("amount")
                amount_per_person = request.POST.get("amount/member")
                receiverID = User.objects.filter(aadhar = receiverAadhar).first()
                if receiverID:
                    CharityAlert.objects.create(
                        text_area = alert,
                        total_amount = totalAmount,
                        receiver = receiverID,
                        receiverNominee = nominee_aadhar,
                        bank_name = bank,
                        upi_id = upi,
                        amount_per_member = amount_per_person,
                        generated_by = request.user,
                        date = datetime.now().date()
                        ).save()
                    return redirect("viewAlert")
                else:
                    return render(request,"userPanel/genAlert.html",{"error":"Receiver Doesn't Exist"})
            else:
                return render(request,"userPanel/genAlert.html")
        else:
            return redirect("home")
    except Exception as e:
        return render(request,"errorPage.html",{"error": e})




# Gallery Management
@login_required
def gallery_management(request):

    if request.method == 'POST':
        user = request.user
        caption = request.POST.get('post-title')
        media_date = request.POST.get("post-date")
        # Post Creation
        new_post = NGOPost.objects.create(
            caption=caption,
            created_by=user,
            event_date=media_date
        )
        media_files = request.FILES.getlist('post-media')
        
        for media in media_files:
            extension = media.name.split('.')[-1].lower()
            media_type = 'image' if extension in ['jpg', 'jpeg', 'png'] else 'video'
            NGOMedia.objects.create(
                post=new_post,
                uploaded_by=user,
                media_file=media,
                media_type=media_type,
                uploaded_at=media_date,
                caption=caption
            )
        # Handle YouTube links
        youtube_links = request.POST.getlist('youtube-link[]')
        for link in youtube_links:
            if not link.strip():
                continue
            NGOMedia.objects.create(
                post=new_post,
                uploaded_by=user,
                media_file=link,   # Save embed URL instead of file
                media_type='video',
                caption=caption
            )
        return redirect("GalleryManagement")
        

    posts = NGOPost.objects.prefetch_related('media_files').order_by('-event_date')
    return render(request,"Gallery/galleryManagement.html",{"posts":posts})

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse   
import json
def delete_post(request,post_id):
    if request.method == 'POST':
        post = NGOPost(id=post_id, created_by=request.user)
        caption = post.caption
        post.delete()
        return JsonResponse({
            'success': True,
            'message': f'Post "{caption}" and its media deleted successfully.'
        })
    return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=405)

@csrf_exempt
def update_designation(request, member_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            designation = data.get("designation")
            print(designation)
            member = User.objects.get(id=member_id)  # or NGOMember
            member.designation = designation
            member.save()

            return JsonResponse({"success": True, "designation": member.designation})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)

    return JsonResponse({"success": False, "error": "Invalid request"}, status=405)

@login_required
def make_home_flag(request, user_id):
    if request.method == "POST":
        user = User.objects.get(id=user_id)
        user.home_flag = not user.home_flag  # toggle
        user.save()
        return JsonResponse({"status": "success", "home_flag": user.home_flag})
    return JsonResponse({"status": "error"}, status=400)


@login_required
def make_admin(request, user_id):
    if request.method == "POST":
        user = User.objects.get(id=user_id)
        user.is_superuser = not user.is_superuser  # toggle
        user.is_staff = user.is_superuser  # keep consistent
        user.save()
        return JsonResponse({"status": "success", "is_superuser": user.is_superuser})
    return JsonResponse({"status": "error"}, status=400)